"""_summary_"""
import os
import sys
from plistlib import InvalidFileException
import time

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

if getattr(sys, 'frozen', False):
    os.chdir(os.path.dirname(sys.executable))
else:
    os.chdir(os.path.dirname(os.path.abspath(__file__)))


def make_choise(customer_id, input_index, column, output_file, value, matching_row, fill):
    """_summary_"""
    print(f"""
          \nFor the customer with ID {customer_id}
          in row №{input_index + 2} of the customer file,
          a value {value} was found that does not match the value
          in your file in row № {matching_row.index[0] + 5}\n""")

    cell = matching_row[column]
    cell.value = value
    matching_index = matching_row.index[0] + 4
    column = matching_row.columns.get_loc(column)

    try:
        wb = load_workbook(filename=output_file, data_only=False)
        ws = wb.active
        ws.cell(row=matching_index + 1, column=column +
                                               1, value=value).fill = fill
        wb.save(output_file)
        print("Data successfully written.")
    except Exception as ex:
        print(f"An error occurred while writing data: {ex}.")
    finally:
        if 'wb' in locals():
            wb.close()


def create_new_row(output_file, kundennummer, fw_seriennummer,
                   fw_mac, l_nummer, name, strasse, plz, ort,
                   fill_green):
    """_summary_"""
    print(f"""
              \nNew data found, which is not in your table!\n""")
    try:
        wb = load_workbook(filename=output_file, data_only=False)
        ws = wb.active
        new_row = ['', '', kundennummer, name, strasse,
                   plz, ort, fw_seriennummer,
                   fw_mac, l_nummer]
        ws.append(new_row)
        last_row_index = ws.max_row
        for col_index, value in enumerate(new_row, start=1):
            cell = ws.cell(row=last_row_index, column=col_index)
            cell.fill = fill_green
        wb.save(output_file)
        print("Data successfully written.")
    except Exception as ex:
        print(f"An error occurred while writing data: {ex}.")
    finally:
        if 'wb' in locals():
            wb.close()


def check_uniqueness(output_file, column_name, fill_green, fill_red):
    """Проверяет уникальность значений в указанном столбце и помечает дубликаты голубым цветом."""
    print(
        f"The process of checking the uniqueness of values in the column {column_name} begins.")
    time.sleep(2)
    fill_blue = PatternFill(start_color="0000FF",
                            end_color="0000FF",
                            fill_type="solid")
    fill_pink = PatternFill(start_color="FFC0CB",
                            end_color="FFC0CB",
                            fill_type="solid")
    try:
        wb = load_workbook(filename=output_file, data_only=False)
        ws = wb.active
        column_index = None
        for column in ws.iter_cols():
            if column[3].value == column_name:
                column_index = column[0].column
                break
        if column_index is None:
            print(f"Column '{column_name}' not found.")
            wb.close()
            return

        checked_column_index = None
        for col in ws.iter_cols():
            if col[3].value == 'Checked':
                checked_column_index = col[0].column
                break

        if checked_column_index is None:
            print("Column 'Checked' not found.")
            wb.close()
            return

        column_values = [cell.value for cell in ws.iter_rows(
            min_row=4, min_col=column_index, max_col=column_index) for cell in cell]
        unique_values = set(column_values)

        for value in unique_values:
            if value in ["not planned", "nicht geplant", None, "-"]:
                continue

            if column_values.count(value) > 1:
                for row in ws.iter_rows(min_row=4, min_col=column_index, max_col=column_index):
                    if row[0].value == value:
                        checked_cell = ws.cell(row=row[0].row, column=checked_column_index)
                        if checked_cell.value != 'checked':
                            checked_cell.value = 'checked'
                            for cell in row:
                                if cell.fill == fill_red or cell.fill == fill_green:
                                    cell.fill = fill_pink
                                else:
                                    cell.fill = fill_blue

        wb.save(output_file)
        print("Found identical and unchecked data, they have been marked!")
    except Exception as ex:
        print(f"An error occurred while writing data: {ex}.")
    finally:
        if 'wb' in locals():
            wb.close()


def check_files(input_file, output_file):
    """_summary_"""
    input_data = pd.read_excel(input_file, dtype={'Kundennummer': str})
    output_data = pd.read_excel(
        output_file, engine='openpyxl', header=3, dtype={'Customer ID': str})
    customer_ids = output_data['Customer ID'].tolist()
    fill_red = PatternFill(start_color="FF0000",
                           end_color="FF0000",
                           fill_type="solid")
    fill_green = PatternFill(start_color="00FF00",
                             end_color="00FF00",
                             fill_type="solid")

    for input_index, input_row in input_data.iterrows():

        if str(input_row['Kundennummer']).isnumeric():
            kundennummer: str = input_row['Kundennummer']
            fw_seriennummer: str = input_row['FW Seriennummer:']
            fw_mac: str = input_row['FW MAC:']
            l_nummer: str = input_row['LNummer:']
            name: str = input_row['Name']
            strasse: str = input_row['Strasse']
            plz: str = input_row['PLZ']
            ort: str = input_row['Ort']
            if kundennummer in customer_ids:
                matching_row: str = output_data[output_data['Customer ID']
                                                == kundennummer]
                customer_id: str = matching_row['Customer ID'].iloc[0]
                name_output: str = matching_row['Name'].iloc[0]
                street: str = matching_row['Street'].iloc[0]
                plz_output: str = matching_row['PLZ'].iloc[0]
                place: str = matching_row['Place'].iloc[0]
                fw_serial_number: str = matching_row['FW Serial number'].iloc[0]
                fw_mac_output: str = matching_row['FW MAC:'].iloc[0]
                l_nummer_output: str = matching_row['LNummer:'].iloc[0]

                if fw_seriennummer != fw_serial_number and fw_seriennummer != 'nicht geplant':
                    make_choise(customer_id, input_index,
                                'FW Serial number',
                                output_file,
                                fw_seriennummer,
                                matching_row, fill_red)

                if fw_mac != fw_mac_output and fw_mac != 'nicht geplant':
                    make_choise(customer_id, input_index,
                                'FW MAC:',
                                output_file,
                                fw_mac,
                                matching_row, fill_red)

                if l_nummer != l_nummer_output and l_nummer != "-" and not pd.isna(l_nummer):
                    make_choise(customer_id, input_index,
                                'LNummer:',
                                output_file,
                                l_nummer,
                                matching_row, fill_red)

                if name != name_output:
                    make_choise(customer_id, input_index,
                                'Name',
                                output_file,
                                name,
                                matching_row, fill_red)

                if strasse != street and strasse != "-" and not pd.isna(strasse):
                    make_choise(customer_id, input_index,
                                'Street',
                                output_file,
                                strasse,
                                matching_row, fill_red)

                if plz != plz_output and plz != "-" and not pd.isna(plz):
                    make_choise(customer_id, input_index,
                                'PLZ',
                                output_file,
                                plz,
                                matching_row, fill_red)

                if ort != place and ort != "-" and not pd.isna(ort):
                    make_choise(customer_id, input_index,
                                'Place',
                                output_file,
                                ort,
                                matching_row, fill_red)

            else:
                create_new_row(output_file, kundennummer,
                               fw_seriennummer, fw_mac, l_nummer,
                               name, strasse, plz, ort, fill_green)

    check_uniqueness(output_file, 'FW Serial number', fill_green, fill_red)
    check_uniqueness(output_file, 'FW MAC:', fill_green, fill_red)
    check_uniqueness(output_file, 'LNummer:', fill_green, fill_red)


if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file_path = 'input.xlsx'
    output_file_path = 'output.xlsx'
    print(f"Current directory: {os.getcwd()}")
    print(f"Files in current directory: {os.listdir()}")
    time.sleep(10)
    if os.path.exists(input_file_path) and os.path.exists(output_file_path):
        try:
            check_files(input_file_path, output_file_path)
            print("\nPress Enter to finish...")
            while True:
                if input("") == "":
                    break
        except FileNotFoundError as e:
            print(f"File not found: {e}.")
            time.sleep(3)
        except PermissionError as e:
            print(f"You do not have permission to access the file: {e}.")
            time.sleep(3)
        except InvalidFileException as e:
            print(f"Invalid file format: {e}.")
            time.sleep(3)
        except Exception as ex:
            print(f"Exception {ex}")
            time.sleep(5)
    else:
        print("One of the files was not found.")
        print("\nPress Enter to finish...")
        while True:
            if input("") == "":
                break
